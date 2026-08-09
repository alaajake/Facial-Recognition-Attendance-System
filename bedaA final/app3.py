import cv2
import pickle
import os
import face_recognition
import serial
import time
import pyttsx3
from openpyxl import Workbook, load_workbook

# Serial setup to communicate with Arduino
arduino = serial.Serial('COM14', 9600)  # Adjust COM port if needed
time.sleep(2)  # Allow time for Arduino to initialize

# Initialize text-to-speech engine
engine = pyttsx3.init()
engine.setProperty('rate', 150)  # Adjust speaking speed if necessary

# Excel file setup
attendance_file = "attendance.xlsx"
if not os.path.exists(attendance_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Timestamp"])  # Headers
    wb.save(attendance_file)

# Load face encodings and names
data_directory = 'data/'
with open(os.path.join(data_directory, 'face_encodings.pkl'), 'rb') as f:
    face_encodings = pickle.load(f)
with open(os.path.join(data_directory, 'names.pkl'), 'rb') as f:
    names = pickle.load(f)

# Initialize video capture
video = cv2.VideoCapture(0)

# Load and prepare background image
background_image_path = 'background.png'
background = cv2.imread(background_image_path)
if background is None:
    print("Error: Background image could not be loaded.")
    video.release()
    cv2.destroyAllWindows()
    exit()

# Get the width and height of the camera frame
frame_width = int(video.get(cv2.CAP_PROP_FRAME_WIDTH))
frame_height = int(video.get(cv2.CAP_PROP_FRAME_HEIGHT))

# Resize the background to match the screen size
screen_width = 1920
screen_height = 1080
background_resized = cv2.resize(background, (screen_width, screen_height))

# Create a fullscreen window
cv2.namedWindow("Frame", cv2.WND_PROP_FULLSCREEN)
cv2.setWindowProperty("Frame", cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)

attendance_taken = False
recognized_name = ""

def save_attendance(name):
    """Save attendance to the Excel file and send data to Arduino."""
    wb = load_workbook(attendance_file)
    ws = wb.active
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    ws.append([name, timestamp])
    wb.save(attendance_file)
    print(f"Attendance saved: {name}, {timestamp}")

    # Send attendance data to Arduino
    arduino.write(f"{name}\n".encode())
    arduino.write(f"{timestamp}\n".encode())

def open_excel_file():
    """Open the attendance Excel file."""
    os.startfile(attendance_file)  # Works for Windows

while True:
    # Read from the camera
    ret, frame = video.read()
    if not ret:
        print("Error: Unable to read from camera.")
        break

    x_offset = 200
    y_offset = 450

    # Overlay the background
    frame_with_background = background_resized.copy()
    frame_with_background[y_offset:y_offset + frame_height, x_offset:x_offset + frame_width] = frame

    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    face_locations = face_recognition.face_locations(rgb_frame)
    face_encodings_frame = face_recognition.face_encodings(rgb_frame, face_locations)

    for face_location, face_encoding in zip(face_locations, face_encodings_frame):
        matches = face_recognition.compare_faces(face_encodings, face_encoding, tolerance=0.5)
        name = "Unknown"

        if True in matches:
            first_match_index = matches.index(True)
            name = names[first_match_index]
            recognized_name = name

        # Draw rectangle and label on the background frame
        (top, right, bottom, left) = face_location
        top += y_offset
        bottom += y_offset
        left += x_offset
        right += x_offset
        cv2.rectangle(frame_with_background, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(frame_with_background, name, (left, top - 10), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

    cv2.imshow("Frame", frame_with_background)

    # Check for keyboard input
    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):  # Quit the program
        break
    elif key == ord('i') and recognized_name:  # Take attendance
        attendance_taken = True
        save_attendance(recognized_name)
        engine.say(f"Attendance taken for {recognized_name}")
        engine.runAndWait()
        break
    elif key == ord('f'):  # Open the Excel file
        open_excel_file()

    # Check for Arduino input
    if arduino.in_waiting > 0:  # Data available from Arduino
        try:
            button_signal = arduino.readline().decode().strip()
            if button_signal == "button1_pressed":  # Simulate 'q' key
                break
            elif button_signal == "button2_pressed":  # Simulate 'f' key
                open_excel_file()
            elif button_signal == "button3_pressed" and recognized_name:  # Simulate 'i' key
                attendance_taken = True
                save_attendance(recognized_name)
                engine.say(f"Attendance taken for {recognized_name}")
                engine.runAndWait()
        except Exception as e:
            print(f"Error processing Arduino input: {e}")
arduino.write(b"System is Closed\n")
# Release resources
video.release()
cv2.destroyAllWindows()
